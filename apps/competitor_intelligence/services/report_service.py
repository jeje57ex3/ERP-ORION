"""
competitor_intelligence/services/report_service.py
Génération de rapports PDF et Excel pour l'analyse concurrentielle.
Nécessite reportlab (PDF) et openpyxl (Excel).
"""
import io
from datetime import datetime
from django.utils import timezone


def generate_competitor_excel_report(company, competitor_ids):
    """Génère un rapport Excel de l'analyse concurrentielle."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise ImportError('openpyxl requis : pip install openpyxl')

    from apps.competitor_intelligence.models import Competitor, CompetitorProduct, CompetitorAdvantage
    from apps.competitor_intelligence.services.analysis_service import generate_competitor_score

    wb = openpyxl.Workbook()

    # ── Feuille résumé ──────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = 'Résumé'
    header_fill = PatternFill('solid', fgColor='2563EB')
    header_font = Font(color='FFFFFF', bold=True)

    headers = ['Concurrent', 'Score', 'Produits suivis', 'Avantages', 'Visiteurs estimés/mois', 'Promotions']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    for row, cid in enumerate(competitor_ids, 2):
        try:
            c = Competitor.objects.get(pk=cid, company=company)
        except Competitor.DoesNotExist:
            continue
        traffic = c.traffic_estimates.order_by('-measured_at').first()
        ws.append([
            c.name,
            generate_competitor_score(c),
            c.products.filter(is_active=True).count(),
            c.advantages.count(),
            f'~{traffic.estimated_monthly_visitors:,} (estimé)' if traffic and traffic.estimated_monthly_visitors else '—',
            c.products.filter(is_active=True, old_price__isnull=False).count(),
        ])

    # ── Feuille produits ────────────────────────────────────────────────────────
    ws2 = wb.create_sheet('Produits concurrents')
    ws2.append(['Concurrent', 'Produit', 'Catégorie', 'Prix', 'Ancien prix', 'Remise %', 'Disponibilité'])
    for cid in competitor_ids:
        try:
            c = Competitor.objects.get(pk=cid, company=company)
        except Competitor.DoesNotExist:
            continue
        for p in c.products.filter(is_active=True).order_by('category', 'name'):
            ws2.append([
                c.name, p.name, p.category,
                float(p.price) if p.price else '',
                float(p.old_price) if p.old_price else '',
                float(p.discount_percent) if p.discount_percent else '',
                p.get_availability_display(),
            ])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def generate_competitor_pdf_report(company, competitor_ids):
    """Génère un rapport PDF de l'analyse concurrentielle."""
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib.units import cm
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
        from reportlab.lib import colors
    except ImportError:
        raise ImportError('reportlab requis : pip install reportlab')

    from apps.competitor_intelligence.models import Competitor
    from apps.competitor_intelligence.services.analysis_service import (
        generate_competitor_score, generate_recommendations,
    )

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=2*cm, rightMargin=2*cm)
    styles = getSampleStyleSheet()
    story = []

    story.append(Paragraph(f'Rapport Analyse Concurrentielle', styles['Title']))
    story.append(Paragraph(f'{company.name} — {timezone.now().strftime("%d/%m/%Y")}', styles['Normal']))
    story.append(Spacer(1, 0.5*cm))

    for cid in competitor_ids:
        try:
            c = Competitor.objects.get(pk=cid, company=company)
        except Competitor.DoesNotExist:
            continue

        story.append(Paragraph(f'Concurrent : {c.name}', styles['Heading2']))
        story.append(Paragraph(f'Score global : {generate_competitor_score(c)}/100', styles['Normal']))

        traffic = c.traffic_estimates.order_by('-measured_at').first()
        if traffic and traffic.estimated_monthly_visitors:
            story.append(Paragraph(
                f'Visiteurs mensuels estimés : ~{traffic.estimated_monthly_visitors:,} '
                f'(source : {traffic.get_source_type_display()}, confiance : {traffic.confidence_score}/10)',
                styles['Normal'],
            ))

        recs = generate_recommendations(company, c)
        if recs:
            story.append(Paragraph('Recommandations :', styles['Heading3']))
            for r in recs:
                story.append(Paragraph(f'• {r}', styles['Normal']))

        story.append(Spacer(1, 0.3*cm))

    doc.build(story)
    buf.seek(0)
    return buf


def generate_market_position_report(company):
    """Génère un rapport de positionnement marché."""
    from apps.competitor_intelligence.services.analysis_service import analyze_market_position
    return analyze_market_position(company)
